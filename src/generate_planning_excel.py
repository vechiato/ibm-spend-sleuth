#!/usr/bin/env python3
"""
IBM Cloud Billing - YAML-to-Excel Planning Generator

This script generates an Excel spreadsheet from a YAML planning configuration,
combining planned vs actual costs from IBM Cloud billing data.

Features:
- Parse YAML planning configuration with groups, months, and filters
- Execute filter commands to get actual billing costs
- Map planned/not_planned status from YAML
- Handle undefined months (mark as not_planned with notes)
- Generate formatted Excel with color coding and summaries
- Create notes sheet with YAML update recommendations

Usage:
    python generate_planning_excel.py --yaml filters.yaml --output planning_report.xlsx
"""

import argparse
import sys
import os
import yaml
import pandas as pd
import subprocess
import re
from datetime import datetime
from typing import Dict, List, Tuple, Any
from dataclasses import dataclass, field
from pathlib import Path

# Import our existing billing parser
try:
    from .ibm_billing_parser import IBMBillingParser
except ImportError:
    from ibm_billing_parser import IBMBillingParser

# Excel imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl package required for Excel generation")
    print("Install with: pip install openpyxl")
    sys.exit(1)

# PyYAML import
try:
    import yaml
except ImportError:
    print("Error: PyYAML package required for YAML parsing")
    print("Install with: pip install PyYAML")
    sys.exit(1)


@dataclass
class GroupConfig:
    """Configuration for a single group from YAML"""
    name: str
    months: Dict[str, Any]  # month -> budget (float) or multi-period budget
    filter_command: str  # For backward compatibility
    filter_commands: List[str] = field(default_factory=list)  # New multi-filter support
    costs: Dict[str, float] = field(default_factory=dict)  # month -> actual cost
    budget_allocations: Dict[str, float] = field(default_factory=dict)  # month -> budget allocation
    planned_costs: Dict[str, float] = field(default_factory=dict)  # month -> planned portion
    not_planned_costs: Dict[str, float] = field(default_factory=dict)  # month -> not_planned portion
    undefined_months: List[str] = field(default_factory=list)


@dataclass
class PlanningData:
    """Complete planning data structure"""
    groups: List[GroupConfig]
    all_months: List[str]
    notes: List[str] = field(default_factory=list)
    total_billing_cost: float = 0.0
    categorized_cost: float = 0.0
    uncategorized_cost: float = 0.0
    coverage_percentage: float = 0.0
    uncategorized_breakdown: Dict[str, Any] = field(default_factory=dict)


class YAMLPlanningParser:
    """Parse YAML planning configuration"""
    
    def __init__(self, yaml_path: str):
        self.yaml_path = yaml_path
        self.planning_data = None
    
    def parse(self) -> PlanningData:
        """Parse YAML file into PlanningData structure"""
        try:
            with open(self.yaml_path, 'r', encoding='utf-8') as file:
                yaml_content = yaml.safe_load(file)
        except FileNotFoundError:
            raise FileNotFoundError(f"YAML file not found: {self.yaml_path}")
        except yaml.YAMLError as e:
            raise ValueError(f"Invalid YAML format: {e}")
        
        if 'groups' not in yaml_content:
            raise ValueError("YAML file must contain 'groups' section")
        
        groups = []
        all_months = set()
        
        for group_data in yaml_content['groups']:
            if 'name' not in group_data:
                raise ValueError("Each group must have a 'name' field")
            if 'filter' not in group_data and 'filters' not in group_data:
                raise ValueError(f"Group '{group_data['name']}' must have either a 'filter' or 'filters' field")
            if 'months' not in group_data:
                raise ValueError(f"Group '{group_data['name']}' must have a 'months' field")
            
            # Handle both old and new filter syntax
            filter_commands = []
            filter_command = ""  # For backward compatibility
            
            if 'filters' in group_data:
                # New array syntax
                if isinstance(group_data['filters'], list):
                    filter_commands = group_data['filters']
                    filter_command = filter_commands[0] if filter_commands else ""
                else:
                    raise ValueError(f"Group '{group_data['name']}' 'filters' field must be a list")
            elif 'filter' in group_data:
                # Old single filter syntax - also check for filter2, filter3, etc.
                filter_command = group_data['filter']
                filter_commands.append(filter_command)
                
                # Check for additional filters (filter2, filter3, etc.)
                filter_num = 2
                while f'filter{filter_num}' in group_data:
                    filter_commands.append(group_data[f'filter{filter_num}'])
                    filter_num += 1
            
            group = GroupConfig(
                name=group_data['name'],
                months=group_data['months'],
                filter_command=filter_command,
                filter_commands=filter_commands
            )
            
            # Process budget allocations
            self._process_budget_allocations(group)
            
            groups.append(group)
            
            # Add all months from budget allocations to the set
            all_months.update(group.budget_allocations.keys())
        
        # Sort months chronologically
        sorted_months = self._sort_months(list(all_months))
        
        self.planning_data = PlanningData(groups=groups, all_months=sorted_months)
        return self.planning_data
    
    def _process_budget_allocations(self, group: GroupConfig):
        """Process budget allocations from YAML months configuration"""
        # Handle case where months is None or empty
        if not group.months:
            group.months = {}
            
        for period, budget_value in group.months.items():
            if isinstance(budget_value, (int, float)) and any(period.startswith(prefix) for prefix in ['Q', 'H', 'Annual-', 'Year-']):
                # Multi-period budget (Q1-25, H1-25, Annual-25)
                months = self._expand_multi_period(period)
                monthly_budget = float(budget_value) / len(months)
                for month in months:
                    group.budget_allocations[month] = monthly_budget
            elif isinstance(budget_value, (int, float)):
                # Single month budget
                group.budget_allocations[period] = float(budget_value)
            elif isinstance(budget_value, str) and budget_value.lower() in ['planned', 'not_planned']:
                # Legacy format - treat as unlimited budget for planned, zero for not_planned
                if budget_value.lower() == 'planned':
                    group.budget_allocations[period] = float('inf')  # Unlimited
                else:
                    group.budget_allocations[period] = 0.0
            else:
                # Unknown format - treat as zero budget
                group.budget_allocations[period] = 0.0
    
    def _expand_multi_period(self, period: str) -> List[str]:
        """Expand multi-period budget into individual months"""
        if period.startswith('Q1-'):
            year = period.split('-')[1]
            return [f'Jan-{year}', f'Feb-{year}', f'Mar-{year}']
        elif period.startswith('Q2-'):
            year = period.split('-')[1]
            return [f'Apr-{year}', f'May-{year}', f'Jun-{year}']
        elif period.startswith('Q3-'):
            year = period.split('-')[1]
            return [f'Jul-{year}', f'Aug-{year}', f'Sep-{year}']
        elif period.startswith('Q4-'):
            year = period.split('-')[1]
            return [f'Oct-{year}', f'Nov-{year}', f'Dec-{year}']
        elif period.startswith('H1-'):
            year = period.split('-')[1]
            return [f'Jan-{year}', f'Feb-{year}', f'Mar-{year}', f'Apr-{year}', f'May-{year}', f'Jun-{year}']
        elif period.startswith('H2-'):
            year = period.split('-')[1]
            return [f'Jul-{year}', f'Aug-{year}', f'Sep-{year}', f'Oct-{year}', f'Nov-{year}', f'Dec-{year}']
        elif period.startswith('Annual-') or period.startswith('Year-'):
            year = period.split('-')[1]
            return [f'Jan-{year}', f'Feb-{year}', f'Mar-{year}', f'Apr-{year}', f'May-{year}', f'Jun-{year}',
                    f'Jul-{year}', f'Aug-{year}', f'Sep-{year}', f'Oct-{year}', f'Nov-{year}', f'Dec-{year}']
        else:
            # Single month - return as is
            return [period]
        
        self.planning_data = PlanningData(groups=groups, all_months=sorted_months)
        return self.planning_data
    
    def _sort_months(self, months: List[str]) -> List[str]:
        """Sort months chronologically (e.g., Jan-25, Feb-25, etc.)"""
        # Create a mapping for month names
        month_order = {
            'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
            'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
        }
        
        def sort_key(month_str):
            try:
                # Parse "Jan-25" format
                month_part, year_part = month_str.split('-')
                year = int(year_part)
                month_num = month_order.get(month_part, 0)
                return (year, month_num)
            except:
                # Fallback: return large number to put at end
                return (9999, 13)
        
        return sorted(months, key=sort_key)


class FilterExecutor:
    """Execute filter commands and collect cost data"""
    
    def __init__(self, data_directory: str = "."):
        self.parser = IBMBillingParser(data_directory)
        self.billing_data = None
        self.last_matched_records = set()  # Track matched record indices
        
        # Create month format mapping
        self.month_mapping = {
            '2025-01': 'Jan-25', '2025-02': 'Feb-25', '2025-03': 'Mar-25',
            '2025-04': 'Apr-25', '2025-05': 'May-25', '2025-06': 'Jun-25',
            '2025-07': 'Jul-25', '2025-08': 'Aug-25', '2025-09': 'Sep-25',
            '2025-10': 'Oct-25', '2025-11': 'Nov-25', '2025-12': 'Dec-25'
        }
    
    def load_billing_data(self):
        """Load all billing data once"""
        print("Loading billing data...")
        self.billing_df = self.parser.load_all_data()  # This returns a DataFrame
        # Keep billing_data as list for compatibility (if needed)
        self.billing_data = self.billing_df.to_dict('records') if not self.billing_df.empty else []
        print(f"Loaded {len(self.billing_data):,} billing records")
    
    def execute_group_filter(self, group: GroupConfig) -> Dict[str, float]:
        """Execute a group's filter commands (supports multiple filters with OR logic)"""
        print(f"Processing group: {group.name}")
        
        if not group.filter_commands:
            print(f"  No filters defined for group {group.name}")
            return {}
        
        try:
            combined_monthly_costs = {}
            all_matched_indices = set()
            
            # Collect all include filter results first, respecting individual month filters
            include_results = []
            exclude_results = []
            
            # Process each filter in the group
            for i, filter_command in enumerate(group.filter_commands):
                print(f"  Executing filter {i+1}/{len(group.filter_commands)}")
                
                # Parse the filter command to extract parameters
                filter_params = self._parse_filter_command(filter_command)
                
                # Add months filter to the filters dict if specified
                if filter_params['months']:
                    filter_params['filters']['Billing Month'] = filter_params['months']
                
                # Execute the filter using our existing parser
                analysis = self.parser.get_filtered_analysis(filter_params['filters'], logic=filter_params['logic'])
                
                if 'filtered_data' in analysis and analysis['filtered_data'] is not None and not analysis['filtered_data'].empty:
                    if filter_params['exclude']:
                        # Store exclude results for later subtraction
                        exclude_results.append(analysis['filtered_data'])
                        print(f"    Filter {i+1} found {len(analysis['filtered_data'])} records to exclude")
                    else:
                        # Store include results
                        include_results.append(analysis['filtered_data'])
                        current_matched = set(analysis['filtered_data'].index)
                        all_matched_indices.update(current_matched)
                        print(f"    Filter {i+1} matched {len(current_matched)} records")
                else:
                    action = "excluded" if filter_params['exclude'] else "matched"
                    print(f"    Filter {i+1} {action} 0 records")
            
            # Combine all include filter results (union/OR logic)
            if include_results:
                # Concatenate all include results and remove duplicates
                current_result_data = pd.concat(include_results, ignore_index=True).drop_duplicates().reset_index(drop=True)
                print(f"  Combined include filters: {len(current_result_data)} total records")
            else:
                # No include filters, start with all data
                current_result_data = self.billing_df.copy()
                print(f"  No include filters, starting with all data: {len(current_result_data)} records")
            
            # Apply exclude filters to the combined result
            if exclude_results:
                # Combine all exclude results
                all_exclude_data = pd.concat(exclude_results, ignore_index=True).drop_duplicates().reset_index(drop=True)
                
                # Remove excluded records from the result
                # We need to match on actual data, not just index since indices may not align
                merge_cols = ['Account ID', 'Billing Month', 'Service Name', 'Instance Name', 'Cost']
                available_cols = [col for col in merge_cols if col in current_result_data.columns and col in all_exclude_data.columns]
                
                if available_cols:
                    # Mark rows to exclude
                    current_result_data = current_result_data.merge(
                        all_exclude_data[available_cols].assign(_exclude=True),
                        on=available_cols,
                        how='left'
                    )
                    # Keep only non-excluded rows
                    excluded_count = current_result_data['_exclude'].sum()
                    current_result_data = current_result_data[current_result_data['_exclude'] != True].drop('_exclude', axis=1, errors='ignore')
                    print(f"  Applied exclude filters: removed {excluded_count} records")
                else:
                    print("  Warning: Could not apply exclude filters due to column mismatch")
            
            # Calculate monthly costs from the final result
            if not current_result_data.empty:
                monthly_costs_df = current_result_data.groupby('Billing Month')['Cost'].sum().reset_index()
                for _, row in monthly_costs_df.iterrows():
                    billing_month = row['Billing Month']  # e.g., "2025-01"
                    yaml_month = self.month_mapping.get(billing_month, billing_month)  # Convert to "Jan-25"
                    cost = float(row['Cost'])
                    combined_monthly_costs[yaml_month] = cost
            
            # Store the matched records from final result
            self.last_matched_records = set(current_result_data.index) if not current_result_data.empty else set()
            
            # Store actual costs
            group.costs = combined_monthly_costs
            
            # Calculate planned vs not_planned costs based on budget allocations
            for month, actual_cost in combined_monthly_costs.items():
                budget = group.budget_allocations.get(month, 0.0)
                
                if budget == float('inf'):
                    # Unlimited budget (legacy "planned")
                    group.planned_costs[month] = actual_cost
                    group.not_planned_costs[month] = 0.0
                elif budget > 0:
                    # Budget defined - split cost
                    if actual_cost <= budget:
                        group.planned_costs[month] = actual_cost
                        group.not_planned_costs[month] = 0.0
                    else:
                        group.planned_costs[month] = budget
                        group.not_planned_costs[month] = actual_cost - budget
                else:
                    # No budget or zero budget - everything is not_planned
                    group.planned_costs[month] = 0.0
                    group.not_planned_costs[month] = actual_cost
            
            total_cost = sum(combined_monthly_costs.values())
            total_planned = sum(group.planned_costs.values())
            total_not_planned = sum(group.not_planned_costs.values())
            
            print(f"  Found {len(combined_monthly_costs)} months with data")
            print(f"  Combined from {len(group.filter_commands)} filters")
            print(f"  Total: ${total_cost:,.2f} (Planned: ${total_planned:,.2f}, Not Planned: ${total_not_planned:,.2f})")
            
            return combined_monthly_costs
            
        except Exception as e:
            print(f"  Error executing filter for group {group.name}: {e}")
            self.last_matched_records = set()
            return {}
    
    def _parse_filter_command(self, command: str) -> Dict[str, Any]:
        """Parse filter command string into parameters"""
        # Remove the python script part and parse arguments
        args_part = command.replace('python src/filter_billing.py', '').strip()
        
        # Simple argument parser for our specific format
        filters = {}
        logic = 'and'  # default
        
        # Extract instances
        instances_match = re.search(r'--instances\s+"([^"]+)"', args_part)
        if instances_match:
            instances_str = instances_match.group(1)
            filters['Instance Name'] = [inst.strip() for inst in instances_str.split(',')]
        
        # Extract services (both --services and --service)
        services_match = re.search(r'--services?\s+"([^"]+)"', args_part)
        if services_match:
            services_str = services_match.group(1)
            filters['Service Name'] = [svc.strip() for svc in services_str.split(',')]
        
        # Extract pattern-column and pattern for pattern-based filtering
        pattern_column_match = re.search(r'--pattern-column\s+"([^"]+)"', args_part)
        pattern_match = re.search(r'--pattern\s+"([^"]+)"', args_part)
        
        if pattern_column_match and pattern_match:
            column_name = pattern_column_match.group(1)
            pattern_value = pattern_match.group(1)
            filters[column_name] = pattern_value
        
        # Extract logic
        logic_match = re.search(r'--logic\s+(\w+)', args_part)
        if logic_match:
            logic = logic_match.group(1)
        
        # Extract exclude flag
        exclude = '--exclude' in args_part
        
        # Extract months for month-specific filtering
        months_match = re.search(r'--months\s+([^\s]+)', args_part)
        months_filter = None
        if months_match:
            months_str = months_match.group(1)
            # Handle both single month and comma-separated months
            months_filter = [month.strip() for month in months_str.split(',')]
        
        return {'filters': filters, 'logic': logic, 'months': months_filter, 'exclude': exclude}

    def calculate_uncategorized_costs(self, planning_data: PlanningData, all_matched_records: set) -> PlanningData:
        """
        Calculate uncategorized costs from billing records that weren't matched by any filter.
        Updates planning_data with uncategorized cost information.
        """
        print("\n📊 Calculating uncategorized costs...")
        
        # We already have billing_df from load_billing_data
        if not hasattr(self, 'billing_df') or self.billing_df is None or self.billing_df.empty:
            print("  ⚠️ No billing data available")
            return planning_data
        
        # Get all billing record indices
        all_indices = set(self.billing_df.index)
        
        # Find unmatched indices
        uncategorized_indices = all_indices - all_matched_records
        
        print(f"  Total billing records: {len(all_indices):,}")
        print(f"  Categorized records: {len(all_matched_records):,}")
        print(f"  Uncategorized records: {len(uncategorized_indices):,}")
        
        # Update planning data with totals
        planning_data.total_billing_cost = float(self.billing_df['Cost'].sum()) if 'Cost' in self.billing_df.columns else 0.0
        planning_data.categorized_cost = sum([sum(group.costs.values()) for group in planning_data.groups])
        planning_data.uncategorized_cost = planning_data.total_billing_cost - planning_data.categorized_cost
        
        if planning_data.total_billing_cost > 0:
            planning_data.coverage_percentage = (planning_data.categorized_cost / planning_data.total_billing_cost) * 100
        else:
            planning_data.coverage_percentage = 100.0
        
        if not uncategorized_indices:
            print("  ✅ All costs are categorized!")
            planning_data.uncategorized_breakdown = {}
            return planning_data
        
        # Create DataFrame with uncategorized records
        uncategorized_df = self.billing_df.loc[list(uncategorized_indices)]
        
        # Calculate monthly breakdown for uncategorized records
        uncategorized_breakdown = {}
        
        if 'Billing Month' in uncategorized_df.columns and 'Cost' in uncategorized_df.columns:
            monthly_groups = uncategorized_df.groupby('Billing Month')
            
            for billing_month, group_df in monthly_groups:
                yaml_month = self.month_mapping.get(billing_month, billing_month)
                
                # Create breakdown by service/resource for this month
                breakdown_items = []
                
                # Group by service and resource
                if 'Service Name' in group_df.columns and 'Resource' in group_df.columns:
                    service_groups = group_df.groupby(['Service Name', 'Resource']).agg({
                        'Cost': 'sum',
                        'Usage': 'sum' if 'Usage' in group_df.columns else lambda x: 0
                    }).round(2)
                    
                    for (service, resource), row in service_groups.iterrows():
                        cost = float(row['Cost'])
                        usage = float(row['Usage']) if 'Usage' in row and pd.notna(row['Usage']) else 0
                        breakdown_items.append({
                            'service': str(service),
                            'resource': str(resource),
                            'cost': cost,
                            'usage': usage
                        })
                else:
                    # Fallback if columns don't exist
                    total_cost = float(group_df['Cost'].sum())
                    breakdown_items.append({
                        'service': 'Unknown',
                        'resource': 'Unknown',
                        'cost': total_cost,
                        'usage': 0
                    })
                
                # Sort by cost descending
                breakdown_items.sort(key=lambda x: x['cost'], reverse=True)
                uncategorized_breakdown[yaml_month] = breakdown_items
        
        planning_data.uncategorized_breakdown = uncategorized_breakdown
        
        print(f"  Total uncategorized costs: ${planning_data.uncategorized_cost:,.2f}")
        print(f"  Coverage percentage: {planning_data.coverage_percentage:.1f}%")
        
        if uncategorized_breakdown:
            print("  📋 Top uncategorized services:")
            # Show top services across all months
            all_services = {}
            for month_breakdown in uncategorized_breakdown.values():
                for item in month_breakdown:
                    service_key = f"{item['service']} - {item['resource']}"
                    if service_key not in all_services:
                        all_services[service_key] = 0
                    all_services[service_key] += item['cost']
            
            # Show top 5 services
            top_services = sorted(all_services.items(), key=lambda x: x[1], reverse=True)[:5]
            for service, total_cost in top_services:
                print(f"    • {service}: ${total_cost:,.2f}")
        
        return planning_data


class ExcelGenerator:
    """Generate Excel spreadsheet with planning data"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.setup_styles()
    
    def setup_styles(self):
        """Setup Excel styling"""
        # Color fills
        self.planned_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        self.not_planned_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        self.undefined_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
        
        # Fonts
        self.header_font = Font(bold=True) #, color="FFFFFF")
        self.bold_font = Font(bold=True)
        self.normal_font = Font()
        
        # Header fill
        #self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
        self.header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
        
        # Alignment
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.right_align = Alignment(horizontal="right", vertical="center")
        
        # Borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_excel(self, planning_data: PlanningData, output_path: str):
        """Generate complete Excel workbook"""
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create sheets
        self._create_main_sheet(planning_data)
        self._create_budget_variance_sheet(planning_data)
        self._create_data_completeness_sheet(planning_data)
        self._create_summary_sheet(planning_data)
        self._create_notes_sheet(planning_data)
        
        # Save workbook
        self.workbook.save(output_path)
        print(f"Excel file saved: {output_path}")
    
    def _create_main_sheet(self, planning_data: PlanningData):
        """Create main planning grid sheet"""
        ws = self.workbook.create_sheet("Planning Grid")
        
        # Headers - Two columns per month (Planned/Not Planned)
        headers = ["Group"]
        for month in planning_data.all_months:
            headers.extend([month, ""])  # Month name, then empty for "Not Planned"
        headers.extend(["Total", "Planned", "Not Planned"])
        
        # Set the headers
        col = 1
        ws.cell(row=1, column=col, value="Group").font = self.header_font
        ws.cell(row=1, column=col).fill = self.header_fill
        ws.cell(row=1, column=col).alignment = self.center_align
        ws.cell(row=1, column=col).border = self.thin_border
        col += 1
        
        # Month headers (spanning two columns each)
        for month in planning_data.all_months:
            # Merge cells for month name
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            cell = ws.cell(row=1, column=col, value=month)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            
            # Sub-headers for Planned/Not Planned
            planned_cell = ws.cell(row=2, column=col, value="Planned")
            planned_cell.font = self.header_font
            planned_cell.fill = self.planned_fill
            planned_cell.alignment = self.center_align
            planned_cell.border = self.thin_border
            
            not_planned_cell = ws.cell(row=2, column=col+1, value="Not Planned")
            not_planned_cell.font = self.header_font
            not_planned_cell.fill = self.not_planned_fill
            not_planned_cell.alignment = self.center_align
            not_planned_cell.border = self.thin_border
            
            col += 2
        
        # Total columns
        for header in ["Total", "Planned", "Not Planned"]:
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            col += 1
        
        # Data rows (starting from row 3)
        row_num = 3
        monthly_planned_totals = {month: 0 for month in planning_data.all_months}
        monthly_not_planned_totals = {month: 0 for month in planning_data.all_months}
        
        for group in planning_data.groups:
            # Group name
            ws.cell(row=row_num, column=1, value=group.name).font = self.bold_font
            
            group_total = 0
            planned_total = 0
            not_planned_total = 0
            
            col = 2
            # Monthly data - two columns per month
            for month in planning_data.all_months:
                planned_cost = group.planned_costs.get(month, 0)
                not_planned_cost = group.not_planned_costs.get(month, 0)
                
                # Planned cost column
                planned_cell = ws.cell(row=row_num, column=col, value=planned_cost if planned_cost > 0 else "")
                if planned_cost > 0:
                    planned_cell.number_format = '$#,##0'
                    planned_cell.fill = self.planned_fill  # Green background for planned costs
                    planned_total += planned_cost
                    monthly_planned_totals[month] += planned_cost
                
                # Not planned cost column
                not_planned_cell = ws.cell(row=row_num, column=col+1, value=not_planned_cost if not_planned_cost > 0 else "")
                if not_planned_cost > 0:
                    not_planned_cell.number_format = '$#,##0'
                    not_planned_cell.fill = self.not_planned_fill  # Red background for not planned costs
                    not_planned_total += not_planned_cost
                    monthly_not_planned_totals[month] += not_planned_cost
                
                # Apply borders
                planned_cell.alignment = self.right_align
                planned_cell.border = self.thin_border
                not_planned_cell.alignment = self.right_align
                not_planned_cell.border = self.thin_border
                
                group_total += planned_cost + not_planned_cost
                col += 2
            
            # Totals
            ws.cell(row=row_num, column=col, value=group_total).number_format = '$#,##0'
            ws.cell(row=row_num, column=col+1, value=planned_total).number_format = '$#,##0'
            ws.cell(row=row_num, column=col+2, value=not_planned_total).number_format = '$#,##0'
            
            row_num += 1
        
        # Add blank row before totals
        row_num += 1
        
        # Monthly totals row
        ws.cell(row=row_num, column=1, value="SUM").font = self.bold_font  # Label for totals row
        
        col = 2
        total_all_planned = 0
        total_all_not_planned = 0
        
        for month in planning_data.all_months:
            planned_total = monthly_planned_totals[month]
            not_planned_total = monthly_not_planned_totals[month]
            
            # Planned column
            planned_cell = ws.cell(row=row_num, column=col, value=planned_total if planned_total > 0 else "")
            if planned_total > 0:
                planned_cell.number_format = '$#,##0'
            planned_cell.font = self.bold_font
            planned_cell.fill = self.header_fill
            planned_cell.border = self.thin_border
            planned_cell.alignment = self.right_align
            
            # Not Planned column
            not_planned_cell = ws.cell(row=row_num, column=col+1, value=not_planned_total if not_planned_total > 0 else "")
            if not_planned_total > 0:
                not_planned_cell.number_format = '$#,##0'
            not_planned_cell.font = self.bold_font
            not_planned_cell.fill = self.header_fill
            not_planned_cell.border = self.thin_border
            not_planned_cell.alignment = self.right_align
            
            total_all_planned += planned_total
            total_all_not_planned += not_planned_total
            col += 2
        
        # Grand totals for the empty row
        total_all = total_all_planned + total_all_not_planned
        ws.cell(row=row_num, column=col, value=total_all).number_format = '$#,##0'
        ws.cell(row=row_num, column=col+1, value=total_all_planned).number_format = '$#,##0'
        ws.cell(row=row_num, column=col+2, value=total_all_not_planned).number_format = '$#,##0'
        
        # Move to next row for "Monthly SUM"
        row_num += 1
        
        # Monthly SUM row (shows total per month)
        monthly_sum_cell = ws.cell(row=row_num, column=1, value="Monthly SUM")
        monthly_sum_cell.font = Font(bold=True, color="FFFFFF")
        monthly_sum_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background
        monthly_sum_cell.alignment = self.center_align
        monthly_sum_cell.border = self.thin_border
        
        col = 2
        for month in planning_data.all_months:
            planned_total = monthly_planned_totals[month]
            not_planned_total = monthly_not_planned_totals[month]
            monthly_total = planned_total + not_planned_total
            
            # Merge the two columns for the monthly total
            ws.merge_cells(start_row=row_num, start_column=col, end_row=row_num, end_column=col+1)
            total_cell = ws.cell(row=row_num, column=col, value=monthly_total if monthly_total > 0 else "")
            if monthly_total > 0:
                total_cell.number_format = '$#,##0'
            total_cell.font = Font(bold=True, color="FFFFFF")
            total_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background
            total_cell.alignment = self.center_align
            total_cell.border = self.thin_border
            
            col += 2
        
        # Add validation section
        self._add_validation_section(ws, planning_data, row_num + 3)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_validation_section(self, ws, planning_data: PlanningData, start_row: int):
        """Add validation section comparing CSV totals with filter totals"""
        row = start_row
        
        # Section header
        validation_header = ws.cell(row=row, column=1, value="VALIDATION - Filter Coverage Analysis")
        validation_header.font = Font(size=14, bold=True, color="FFFFFF")
        validation_header.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")  # Orange background
        validation_header.alignment = self.center_align
        validation_header.border = self.thin_border
        
        # Merge across multiple columns for the header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(planning_data.all_months) * 2 + 3)
        row += 2
        
        try:
            # Initialize billing parser to get total costs from CSV
            parser = IBMBillingParser("data/billing")
            df = parser.load_all_data()
            
            # Calculate monthly totals from CSV
            csv_monthly_totals = {}
            month_mapping = {
                "2025-01": "Jan-25", "2025-02": "Feb-25", "2025-03": "Mar-25",
                "2025-04": "Apr-25", "2025-05": "May-25", "2025-06": "Jun-25",
                "2025-07": "Jul-25", "2025-08": "Aug-25", "2025-09": "Sep-25",
                "2025-10": "Oct-25", "2025-11": "Nov-25", "2025-12": "Dec-25"
            }
            
            for csv_month, display_month in month_mapping.items():
                if display_month in planning_data.all_months:
                    month_data = df[df['Billing Month'] == csv_month]
                    csv_monthly_totals[display_month] = month_data['Cost'].sum()
            
            # Calculate filtered totals for each month
            filtered_monthly_totals = {}
            for month in planning_data.all_months:
                filtered_monthly_totals[month] = sum(
                    group.costs.get(month, 0) for group in planning_data.groups
                )
            
            # Create validation table headers
            validation_headers = ["Month", "Gran total from CSV", "Filtered Total", "Difference"]
            for col, header in enumerate(validation_headers, 1):
                header_cell = ws.cell(row=row, column=col, value=header)
                header_cell.font = self.header_font
                header_cell.fill = self.header_fill
                header_cell.border = self.thin_border
                header_cell.alignment = self.center_align
            row += 1
            
            # Add monthly validation rows
            for month in planning_data.all_months:
                csv_total = csv_monthly_totals.get(month, 0)
                filtered_total = filtered_monthly_totals.get(month, 0)
                difference = filtered_total - csv_total
                
                # Month name
                month_cell = ws.cell(row=row, column=1, value=month)
                month_cell.border = self.thin_border
                month_cell.alignment = self.center_align
                
                # CSV total
                csv_cell = ws.cell(row=row, column=2, value=csv_total)
                csv_cell.number_format = '$#,##0'
                csv_cell.border = self.thin_border
                csv_cell.alignment = self.right_align
                
                # Filtered total
                filtered_cell = ws.cell(row=row, column=3, value=filtered_total)
                filtered_cell.number_format = '$#,##0'
                filtered_cell.border = self.thin_border
                filtered_cell.alignment = self.right_align
                
                # Difference with color coding
                diff_cell = ws.cell(row=row, column=4, value=difference)
                diff_cell.number_format = '$#,##0'
                diff_cell.border = self.thin_border
                diff_cell.alignment = self.right_align
                
                # Color code the difference with $1 tolerance
                if difference > 1:
                    # Positive difference > $1 (filtered > CSV) = Red (overlapping filters)
                    diff_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    diff_cell.font = Font(color="CC0000", bold=True)
                elif difference < -2:
                    # Negative difference < -$2 (filtered < CSV) = Yellow (missing coverage)
                    diff_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    diff_cell.font = Font(color="CC6600", bold=True)
                else:
                    # Within $1 tolerance = Green (acceptable match)
                    diff_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    diff_cell.font = Font(color="006600", bold=True)
                
                row += 1
            
            # Add explanation
            row += 1
            explanation_text = [
                "🔴 Red (>+$1): Filters are overlapping - same costs counted multiple times",
                "🟡 Yellow (<-$1): Filters need improvement - some costs not captured",
                "🟢 Green (±$1): Acceptable match - within $1 tolerance"
            ]
            
            for text in explanation_text:
                explanation_cell = ws.cell(row=row, column=1, value=text)
                explanation_cell.font = Font(size=10, italic=True)
                if "Red" in text:
                    explanation_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif "Yellow" in text:
                    explanation_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif "Green" in text:
                    explanation_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                row += 1
                
        except Exception as e:
            # If validation fails, add error message
            error_cell = ws.cell(row=row, column=1, value=f"Validation failed: {str(e)}")
            error_cell.font = Font(color="CC0000", italic=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    def _create_budget_variance_sheet(self, planning_data: PlanningData):
        """Create budget variance analysis sheet"""
        ws = self.workbook.create_sheet("Budget Variance")
        
        # Headers
        headers = ["Group", "Month", "Budget", "Actual", "Variance", "Variance %", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # Data rows
        row_num = 2
        for group in planning_data.groups:
            for month in planning_data.all_months:
                budget = group.budget_allocations.get(month, 0)
                actual = group.costs.get(month, 0)
                
                if budget > 0 or actual > 0:  # Only show months with budget or actual costs
                    # Handle unlimited budget display
                    budget_display = "Unlimited" if budget == float('inf') else budget
                    variance = 0 if budget == float('inf') else actual - budget
                    variance_pct = 0 if budget == float('inf') or budget == 0 else (variance / budget) * 100
                    
                    # Status determination
                    if budget == float('inf'):
                        status = "Planned"
                    elif budget == 0:
                        status = "Not Planned"
                    elif actual <= budget:
                        status = "Within Budget"
                    else:
                        status = "Over Budget"
                    
                    # Fill row data
                    ws.cell(row=row_num, column=1, value=group.name)
                    ws.cell(row=row_num, column=2, value=month)
                    
                    budget_cell = ws.cell(row=row_num, column=3, value=budget_display)
                    if isinstance(budget_display, (int, float)):
                        budget_cell.number_format = '$#,##0'
                    
                    actual_cell = ws.cell(row=row_num, column=4, value=actual if actual > 0 else "")
                    if actual > 0:
                        actual_cell.number_format = '$#,##0'
                    
                    variance_cell = ws.cell(row=row_num, column=5, value=variance if budget != float('inf') else "")
                    if budget != float('inf') and variance != 0:
                        variance_cell.number_format = '$#,##0'
                        if variance > 0:
                            variance_cell.fill = self.not_planned_fill  # Red for over budget
                        else:
                            variance_cell.fill = self.planned_fill  # Green for under budget
                    
                    variance_pct_cell = ws.cell(row=row_num, column=6, value=variance_pct if budget != float('inf') else "")
                    if budget != float('inf') and variance_pct != 0:
                        variance_pct_cell.number_format = '0.0%'
                        if variance_pct > 0:
                            variance_pct_cell.fill = self.not_planned_fill
                        else:
                            variance_pct_cell.fill = self.planned_fill
                    
                    status_cell = ws.cell(row=row_num, column=7, value=status)
                    if status == "Over Budget":
                        status_cell.fill = self.not_planned_fill
                    elif status == "Within Budget":
                        status_cell.fill = self.planned_fill
                    elif status == "Not Planned":
                        status_cell.fill = self.undefined_fill
                    
                    # Apply borders
                    for col in range(1, 8):
                        ws.cell(row=row_num, column=col).border = self.thin_border
                        ws.cell(row=row_num, column=col).alignment = self.center_align
                    
                    row_num += 1
        
        # Auto-fit columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_summary_sheet(self, planning_data: PlanningData):
        """Create summary analysis sheet"""
        ws = self.workbook.create_sheet("Summary")
        
        row = 1
        ws.cell(row=row, column=1, value="IBM Cloud Billing - Planning Summary").font = Font(size=16, bold=True)
        row += 3
        
        # Group summaries
        ws.cell(row=row, column=1, value="Group Summaries").font = self.bold_font
        row += 1
        
        headers = ["Group", "Total Cost", "Planned Cost", "Not Planned Cost", "Months with Data"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1
        
        for group in planning_data.groups:
            total_cost = sum(group.costs.values())
            planned_cost = sum(cost for month, cost in group.costs.items() 
                             if month in group.months and group.months[month] == "planned")
            not_planned_cost = total_cost - planned_cost
            months_count = len([m for m in group.costs.keys() if group.costs[m] > 0])
            
            ws.cell(row=row, column=1, value=group.name)
            ws.cell(row=row, column=2, value=total_cost).number_format = '$#,##0'
            ws.cell(row=row, column=3, value=planned_cost).number_format = '$#,##0'
            ws.cell(row=row, column=4, value=not_planned_cost).number_format = '$#,##0'
            ws.cell(row=row, column=5, value=months_count)
            row += 1

    def _create_data_completeness_sheet(self, planning_data: PlanningData):
        """Create data completeness analysis sheet"""
        ws = self.workbook.create_sheet("Data Completeness")
        
        # Set up special styling for uncategorized content
        uncategorized_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
        warning_font = Font(bold=True, color="FF8C00")  # Dark orange
        
        row = 1
        ws.cell(row=row, column=1, value="📊 DATA COMPLETENESS ANALYSIS").font = Font(size=16, bold=True)
        row += 2
        
        # Summary section
        ws.cell(row=row, column=1, value="Coverage Summary:").font = self.bold_font
        row += 1
        
        total_cost = getattr(planning_data, 'total_billing_cost', 0.0)
        categorized_cost = getattr(planning_data, 'categorized_cost', 0.0)
        uncategorized_cost = getattr(planning_data, 'uncategorized_cost', 0.0)
        coverage_percentage = getattr(planning_data, 'coverage_percentage', 0.0)
        
        ws.cell(row=row, column=1, value=f"Total Billing Cost:")
        ws.cell(row=row, column=2, value=total_cost).number_format = '$#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value=f"Categorized Cost:")
        ws.cell(row=row, column=2, value=categorized_cost).number_format = '$#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value=f"⚠️ Uncategorized Cost:")
        cell = ws.cell(row=row, column=2, value=uncategorized_cost)
        cell.number_format = '$#,##0.00'
        if uncategorized_cost > 0:
            cell.fill = uncategorized_fill
            cell.font = warning_font
        row += 1
        
        ws.cell(row=row, column=1, value=f"Coverage Percentage:")
        cell = ws.cell(row=row, column=2, value=coverage_percentage/100)
        cell.number_format = '0.0%'
        if coverage_percentage < 100:
            cell.fill = uncategorized_fill
        row += 2
        
        # Check if we have uncategorized breakdown
        uncategorized_breakdown = getattr(planning_data, 'uncategorized_breakdown', {})
        
        if uncategorized_cost > 0 and uncategorized_breakdown:
            # Uncategorized costs breakdown
            ws.cell(row=row, column=1, value="⚠️ UNCATEGORIZED COSTS BREAKDOWN").font = Font(size=14, bold=True, color="FF8C00")
            row += 2
            
            # Headers
            headers = ["Month", "Service Name", "Resource", "Cost", "Usage"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.thin_border
            row += 1
            
            # Data rows
            for month, breakdown_items in uncategorized_breakdown.items():
                for item in breakdown_items:
                    ws.cell(row=row, column=1, value=month)
                    ws.cell(row=row, column=2, value=item['service'])
                    ws.cell(row=row, column=3, value=item['resource'])
                    
                    cost_cell = ws.cell(row=row, column=4, value=item['cost'])
                    cost_cell.number_format = '$#,##0.00'
                    cost_cell.fill = uncategorized_fill
                    
                    usage_cell = ws.cell(row=row, column=5, value=item.get('usage', 0))
                    usage_cell.number_format = '#,##0.00'
                    
                    row += 1
            
            row += 1
            
            # Monthly totals
            ws.cell(row=row, column=1, value="Monthly Uncategorized Totals:").font = self.bold_font
            row += 1
            
            month_totals = {}
            for month, breakdown_items in uncategorized_breakdown.items():
                month_total = sum(item['cost'] for item in breakdown_items)
                month_totals[month] = month_total
                
                ws.cell(row=row, column=1, value=month)
                cell = ws.cell(row=row, column=2, value=month_total)
                cell.number_format = '$#,##0.00'
                cell.fill = uncategorized_fill
                row += 1
            
        else:
            # All costs are categorized
            ws.cell(row=row, column=1, value="✅ ALL COSTS ARE CATEGORIZED!").font = Font(size=14, bold=True, color="008000")
            row += 1
            ws.cell(row=row, column=1, value="Great job! Your filters capture 100% of the billing data.")
        
        row += 2
        
        # Recommendations section
        ws.cell(row=row, column=1, value="💡 RECOMMENDATIONS").font = Font(size=14, bold=True)
        row += 2
        
        if uncategorized_cost > 0:
            ws.cell(row=row, column=1, value="To improve data completeness:")
            row += 1
            
            ws.cell(row=row, column=1, value="1. Review the uncategorized services above")
            row += 1
            
            ws.cell(row=row, column=1, value="2. Add new filter groups to your YAML configuration")
            row += 1
            
            ws.cell(row=row, column=1, value="3. Use specific service names and instance patterns")
            row += 1
            
            ws.cell(row=row, column=1, value="4. Test your updated filters and regenerate the report")
            row += 1
        else:
            ws.cell(row=row, column=1, value="✅ Your filter configuration is complete!")
            row += 1
            ws.cell(row=row, column=1, value="All billing data is properly categorized.")
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
    
    def _create_notes_sheet(self, planning_data: PlanningData):
        """Create notes and recommendations sheet"""
        ws = self.workbook.create_sheet("Notes & Recommendations")
        
        row = 1
        ws.cell(row=row, column=1, value="YAML Update Recommendations").font = Font(size=16, bold=True)
        row += 2
        
        # Color legend
        ws.cell(row=row, column=1, value="Color Legend:").font = self.bold_font
        row += 1
        
        # Green - Planned
        cell = ws.cell(row=row, column=1, value="🟢 Green = Planned months (costs included in planning)")
        cell.fill = self.planned_fill
        row += 1
        
        # Red - Not Planned
        cell = ws.cell(row=row, column=1, value="🔴 Red = Not planned months (costs marked as unplanned)")
        cell.fill = self.not_planned_fill
        row += 1
        
        # Yellow - Undefined
        cell = ws.cell(row=row, column=1, value="🟡 Yellow = Undefined months (found in data but not in YAML)")
        cell.fill = self.undefined_fill
        row += 1
        
        # White - No cost
        ws.cell(row=row, column=1, value="⚪ White = No cost data for the month")
        row += 3
        
        ws.cell(row=row, column=1, value="The following months were found in your billing data but are not defined in filters.yaml.")
        row += 1
        ws.cell(row=row, column=1, value="They have been included as 'not_planned' by default.")
        row += 2
        
        has_undefined = False
        for group in planning_data.groups:
            undefined_months = []
            for month, cost in group.costs.items():
                if month not in group.months and cost > 0:
                    undefined_months.append((month, cost))
            
            if undefined_months:
                has_undefined = True
                ws.cell(row=row, column=1, value=f"Group \"{group.name}\":").font = self.bold_font
                row += 1
                
                for month, cost in undefined_months:
                    ws.cell(row=row, column=1, value=f"  - {month}: not_planned  # Currently showing ${cost:,.2f}")
                    row += 1
                row += 1
        
        if not has_undefined:
            ws.cell(row=row, column=1, value="✅ All months in billing data are properly defined in your YAML file!")
        else:
            row += 1
            ws.cell(row=row, column=1, value="Action: Edit filters.yaml and change 'not_planned' to 'planned' if desired.").font = self.bold_font
        
        # Set column width
        ws.column_dimensions['A'].width = 80


def main():
    parser = argparse.ArgumentParser(description='Generate Excel planning report from YAML configuration')
    parser.add_argument('--yaml', required=True, help='Path to YAML planning configuration file')
    parser.add_argument('--output', required=True, help='Output Excel file path')
    parser.add_argument('--data-dir', default='data/billing', help='Directory containing billing CSV files (default: data/billing)')
    
    args = parser.parse_args()
    
    # Validate inputs
    if not os.path.exists(args.yaml):
        print(f"Error: YAML file not found: {args.yaml}")
        sys.exit(1)
    
    # Ensure output directory exists
    output_dir = os.path.dirname(args.output)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    try:
        print("🚀 Starting YAML-to-Excel Planning Generator")
        print("=" * 50)
        
        # Parse YAML
        print(f"📋 Parsing YAML configuration: {args.yaml}")
        yaml_parser = YAMLPlanningParser(args.yaml)
        planning_data = yaml_parser.parse()
        print(f"   Found {len(planning_data.groups)} groups")
        
        # Execute filters
        print(f"💰 Executing filter commands...")
        executor = FilterExecutor(args.data_dir)
        executor.load_billing_data()
        
        # Track matched records for uncategorized detection
        all_matched_records = set()
        
        for group in planning_data.groups:
            monthly_costs = executor.execute_group_filter(group)
            group.costs = monthly_costs
            
            # Track which records were matched by this group
            if hasattr(executor, 'last_matched_records'):
                all_matched_records.update(executor.last_matched_records)
        
        # Calculate uncategorized costs
        print(f"🔍 Analyzing data completeness...")
        planning_data = executor.calculate_uncategorized_costs(planning_data, all_matched_records)
        
        # Generate Excel
        print(f"📊 Generating Excel report: {args.output}")
        excel_generator = ExcelGenerator()
        excel_generator.generate_excel(planning_data, args.output)
        
        print("=" * 50)
        print("✅ Planning report generated successfully!")
        print(f"📁 Output file: {args.output}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()